"""
Excel exporter in practical 1C-compatible format.

Creates two sheets:
- "Проверка" for accountant-friendly validation blocks
- "1C_Импорт" as flat rows for import
"""

from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

CHECK_SHEET_NAME = "Проверка"
IMPORT_SHEET_NAME = "1C_Импорт"

IMPORT_COLUMNS_1C = [
    "Дата документа",
    "Номер документа-основания",
    "Продавец",
    "ИНН продавца",
    "Вид документа",
    "Номенклатура",
    "Ед. изм.",
    "Количество",
    "Цена",
    "Сумма",
    "Ставка НДС",
    "Сумма НДС",
    "Комментарий",
]

CHECK_BLOCK_ROWS = [
    "Дата чека",
    "Номер чека",
    "Продавец",
    "ИНН продавца",
    "Итоговая сумма",
    "НДС всего",
    "Количество позиций",
    "Статус проверки",
]


def _auto_fit(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[letter].width = min(width + 2, 50)


def _num(value) -> float | None:
    """Safe cast to float; returns None on failure."""
    if value is None:
        return None


def _parse_iso_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        try:
            return date.fromisoformat(raw)
        except ValueError:
            try:
                return datetime.fromisoformat(raw).date()
            except ValueError:
                return None
    return None


def _is_ready_for_export(doc_date, seller: str, items: list[dict]) -> bool:
    return bool(doc_date and seller and items)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def build_excel_1c(results: list[dict], filename: str = "receipts_1c.xlsx") -> str:
    """
    Формирует Excel-файл в формате для импорта в 1С.

    Args:
        results: список канонических результатов (output of process_receipt).
        filename: путь к выходному файлу.

    Returns:
        Путь к сохранённому файлу.
    """
    wb = Workbook()
    ws_check = wb.active
    ws_check.title = CHECK_SHEET_NAME
    ws_import = wb.create_sheet(IMPORT_SHEET_NAME)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    ws_check["A1"] = "Проверка выгрузки для 1С"
    ws_check["A1"].font = Font(bold=True)
    ws_check["A1"].alignment = Alignment(horizontal="left")

    ws_import.append(IMPORT_COLUMNS_1C)
    for col_idx in range(1, len(IMPORT_COLUMNS_1C) + 1):
        cell = ws_import.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    check_row = 3
    for idx, r in enumerate(results, start=1):
        receipt = r.get("receipt", {}) or {}
        merchant = r.get("merchant", {}) or {}
        items = r.get("items", []) or []
        totals = r.get("totals", {}) or {}
        taxes = r.get("taxes", {}) or {}

        doc_date_str = receipt.get("date", "") or ""
        doc_date = _parse_iso_date(doc_date_str)
        receipt_number = receipt.get("receipt_number", "") or ""
        inn = merchant.get("inn", "") or ""
        seller = merchant.get("organization", "") or ""
        total_sum = _num(totals.get("total"))
        total_vat = _num(taxes.get("total_vat"))

        status = (
            "Готово к выгрузке"
            if _is_ready_for_export(doc_date, seller, items)
            else "Проверить вручную"
        )

        ws_check.cell(row=check_row, column=1, value=f"Чек #{idx}").font = Font(bold=True)
        check_row += 1

        values = [
            doc_date if doc_date else doc_date_str,
            receipt_number,
            seller,
            inn,
            total_sum,
            total_vat,
            len(items),
            status,
        ]

        for label, value in zip(CHECK_BLOCK_ROWS, values):
            ws_check.cell(row=check_row, column=1, value=label).font = header_font
            value_cell = ws_check.cell(row=check_row, column=2, value=value)
            if label == "Дата чека" and isinstance(value, date):
                value_cell.number_format = "DD.MM.YYYY"
            if label in ("Итоговая сумма", "НДС всего") and isinstance(value, (int, float)):
                value_cell.number_format = "0.00"
            check_row += 1

        check_row += 1

        if not items:
            ws_import.append(
                [
                    doc_date if doc_date else doc_date_str,
                    receipt_number,
                    seller,
                    inn,
                    "Кассовый чек",
                    "Нет данных",
                    "шт",
                    None,
                    None,
                    None,
                    "",
                    None,
                    "Распознано из фото чека",
                ]
            )
            continue

        for item in items:
            ws_import.append(
                [
                    doc_date if doc_date else doc_date_str,
                    receipt_number,
                    seller,
                    inn,
                    "Кассовый чек",
                    item.get("name", ""),
                    "шт",
                    _num(item.get("quantity")),
                    _num(item.get("price_per_unit")),
                    _num(item.get("total_price")),
                    item.get("vat_rate", ""),
                    _num(item.get("vat_amount")),
                    "Распознано из фото чека",
                ]
            )

    for row in ws_import.iter_rows(min_row=2, max_row=ws_import.max_row):
        doc_date_cell = row[0]
        qty_cell = row[7]
        price_cell = row[8]
        sum_cell = row[9]
        vat_sum_cell = row[11]

        if isinstance(doc_date_cell.value, date):
            doc_date_cell.number_format = "DD.MM.YYYY"

        qty_value = qty_cell.value
        if isinstance(qty_value, (int, float)):
            is_integer = float(qty_value).is_integer()
            qty_cell.number_format = "0" if is_integer else "0.###"

        for money_cell in (price_cell, sum_cell, vat_sum_cell):
            if isinstance(money_cell.value, (int, float)):
                money_cell.number_format = "0.00"

    _auto_fit(ws_check)
    _auto_fit(ws_import)
    wb.save(filename)
    return filename
