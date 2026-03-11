import sys
import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from src.vision_utils import prepare_image
from src.openai_client import extract_receipt_data_from_image

def process_receipt(image_path):
    """Обрабатывает одно изображение через OpenAI Vision API"""
    print(f"\n🔄 Обработка {image_path}...")
    
    # Проверяем изображение
    try:
        prepared_path = prepare_image(image_path)
    except Exception as e:
        print(f"❌ Ошибка при подготовке изображения: {e}")
        return None
    
    # Отправляем напрямую в OpenAI Vision
    data = extract_receipt_data_from_image(prepared_path)
    return data

def auto_fit_columns(worksheet):
    """Автоматически подбирает ширину колонок по содержимому"""
    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def save_to_excel(results, filename="receipts.xlsx"):
    """Сохраняет результаты в Excel-файл с двумя листами: сводка и товары."""
    wb = Workbook()
    
    # --- Лист 1: Сводка по чекам ---
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    
    # Заголовки
    summary_headers = ["№", "Организация", "ИНН", "Дата", "Общая сумма", "НДС всего"]
    ws_summary.append(summary_headers)
    
    # Форматирование заголовков
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Данные сводки
    for idx, r in enumerate(results, 1):
        row = [
            idx,
            r.get("organization", ""),
            r.get("inn", ""),
            r.get("date", ""),
            r.get("total", ""),
            r.get("total_vat", "")
        ]
        ws_summary.append(row)
    
    # Автоширина колонок на листе сводки
    for col in range(1, len(summary_headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws_summary.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        auto_fit_columns(ws_summary)

    
    # --- Лист 2: Товары (детализация) ---
    auto_fit_columns(ws_items)

    
    # Заголовки для товаров
    items_headers = ["№ чека", "Организация", "ИНН", "Дата", 
                     "Наименование товара", "Цена за ед.", "Кол-во", 
                     "Стоимость", "Ставка НДС", "Сумма НДС"]
    ws_items.append(items_headers)
    
    # Форматирование заголовков
    for col, header in enumerate(items_headers, 1):
        cell = ws_items.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Данные товаров
    for idx, r in enumerate(results, 1):
        organization = r.get("organization", "")
        inn = r.get("inn", "")
        date = r.get("date", "")
        items = r.get("items", [])
        if not items:
            ws_items.append([idx, organization, inn, date, "Нет данных", "", "", "", "", ""])
        else:
            for item in items:
                row = [
                    idx,
                    organization,
                    inn,
                    date,
                    item.get("name", ""),
                    item.get("price_per_unit", ""),
                    item.get("quantity", ""),
                    item.get("total_price", ""),
                    item.get("vat_rate", ""),
                    item.get("vat_amount", "")
                ]
                ws_items.append(row)
    
    # Автоширина колонок на листе товаров
    for col in range(1, len(items_headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws_items.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        ws_items.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    wb.save(filename)
    print(f"\n✅ Результаты сохранены в файл {filename}")

def main():
    if len(sys.argv) < 2:
        print("Использование: python main.py <путь_к_изображению_или_папке>")
        return
    
    path = sys.argv[1]
    
    if os.path.isfile(path):
        files = [path]
    else:
        files = [os.path.join(path, f) for f in os.listdir(path) 
                 if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
    
    if not files:
        print(f"❌ Не найдено изображений (.jpg, .jpeg, .png) по пути: {path}")
        print("💡 Убедитесь, что папка содержит файлы с расширением .jpg, .jpeg или .png")
        return

    
    results = []
    for f in files:
        res = process_receipt(f)
        if res:
            results.append(res)
            print(f"✅ Чек обработан успешно")
        else:
            print(f"❌ Не удалось обработать чек")
    
    if not results:
        print("❌ Нет данных для сохранения.")
        return
    
    # Вывод в консоль
    for i, r in enumerate(results, 1):
        print(f"\n📄 --- Чек {i} ---")
        print(f"Организация: {r.get('organization')}")
        print(f"ИНН: {r.get('inn')}")
        print(f"Дата: {r.get('date')}")
        print(f"Всего: {r.get('total')} руб.")
        print(f"НДС всего: {r.get('total_vat')} руб.")
        print("Товары:")
        for item in r.get('items', []):
            name = item.get('name', '')
            price = item.get('total_price', '')
            print(f"  {name}: {price} руб.")
    
    # Сохраняем в Excel
    save_to_excel(results, "receipts.xlsx")

if __name__ == "__main__":
    main()