"""
Core receipt processing functions.

This module is the single entry point for receipt parsing and Excel export.
Both CLI (main.py) and API (api/app.py) import from here.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from src.vision_utils import prepare_image
from src.openai_client import extract_receipt_data_from_image


def process_receipt(image_path: str) -> dict | None:
    """Обрабатывает одно изображение через pipeline и возвращает канонический результат."""
    print(f"\n🔄 Обработка {image_path}...")

    try:
        prepared_path = prepare_image(image_path)
    except Exception as e:
        print(f"❌ Ошибка при подготовке изображения: {e}")
        return None

    return extract_receipt_data_from_image(prepared_path)


def auto_fit_columns(worksheet) -> None:
    """Автоматически подбирает ширину колонок по содержимому."""
    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)


def save_to_excel(results: list[dict], filename: str = "receipts.xlsx") -> str:
    """
    Сохраняет результаты в Excel-файл с двумя листами: сводка и товары.

    Returns:
        Путь к сохранённому файлу.
    """
    wb = Workbook()

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # --- Лист 1: Сводка по чекам ---
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    summary_headers = ["№", "Номер чека", "Организация", "ИНН", "Дата", "Общая сумма", "НДС всего"]
    ws_summary.append(summary_headers)

    for col, _ in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, r in enumerate(results, 1):
        receipt = r.get("receipt", {}) or {}
        merchant = r.get("merchant", {}) or {}
        totals = r.get("totals", {}) or {}
        taxes = r.get("taxes", {}) or {}

        ws_summary.append([
            idx,
            receipt.get("receipt_number", ""),
            merchant.get("organization", ""),
            merchant.get("inn", ""),
            receipt.get("date", ""),
            totals.get("total", ""),
            taxes.get("total_vat", ""),
        ])

    auto_fit_columns(ws_summary)

    # --- Лист 2: Товары (детализация) ---
    ws_items = wb.create_sheet("Товары")

    items_headers = [
        "№", "Номер чека", "Организация", "ИНН", "Дата",
        "Наименование товара", "Цена за ед.", "Кол-во",
        "Стоимость", "Ставка НДС", "Сумма НДС",
    ]
    ws_items.append(items_headers)

    for col, _ in enumerate(items_headers, 1):
        cell = ws_items.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, r in enumerate(results, 1):
        receipt = r.get("receipt", {}) or {}
        merchant = r.get("merchant", {}) or {}
        items = r.get("items", []) or []

        organization = merchant.get("organization", "") or ""
        inn = merchant.get("inn", "") or ""
        date = receipt.get("date", "") or ""
        receipt_number = receipt.get("receipt_number", "") or ""

        if not items:
            ws_items.append([idx, receipt_number, organization, inn, date,
                             "Нет данных", "", "", "", "", ""])
        else:
            for item in items:
                ws_items.append([
                    idx,
                    receipt_number,
                    organization,
                    inn,
                    date,
                    item.get("name", ""),
                    item.get("price_per_unit", ""),
                    item.get("quantity", ""),
                    item.get("total_price", ""),
                    item.get("vat_rate", ""),
                    item.get("vat_amount", ""),
                ])

    auto_fit_columns(ws_items)

    wb.save(filename)
    print(f"\n✅ Результаты сохранены в файл {filename}")
    return filename
