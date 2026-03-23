from pathlib import Path

from openpyxl import load_workbook

from api.exporters.csv_1c import build_csv_1c_bytes
from api.exporters.excel_1c import build_excel_1c


EXPECTED_HEADERS = [
    "Дата документа",
    "Номер документа-основания",
    "Продавец",
    "ИНН продавца",
    "Вид документа",
    "Номенклатура",
    "Ед. изм.",
    "Количество",
    "Цена",
    "Сумма",
    "Ставка НДС",
    "Сумма НДС",
    "Комментарий",
]

CHECK_LABELS = {
    "Дата чека",
    "Номер чека",
    "Продавец",
    "ИНН продавца",
    "Итоговая сумма",
    "НДС всего",
    "Количество позиций",
    "Статус проверки",
}


def test_excel_1c_file_created_and_has_expected_structure(tmp_path, sample_receipt_result):
    out_path = tmp_path / "out.xlsx"
    returned = build_excel_1c([sample_receipt_result], str(out_path))

    assert Path(returned).exists()

    wb = load_workbook(out_path)
    assert "Проверка" in wb.sheetnames
    assert "1C_Импорт" in wb.sheetnames
    assert wb.active.title in {"Проверка", "1C_Импорт"}

    ws = wb["1C_Импорт"]

    headers = [cell.value for cell in ws[1]]
    assert headers == EXPECTED_HEADERS
    assert ws.max_row >= 2

    ws_check = wb["Проверка"]
    labels_in_sheet = {
        cell.value
        for row in ws_check.iter_rows(min_row=1, max_row=ws_check.max_row, min_col=1, max_col=1)
        for cell in row
        if isinstance(cell.value, str)
    }
    assert CHECK_LABELS.issubset(labels_in_sheet)


def test_csv_1c_encoding_separator_and_headers(sample_receipt_result):
    csv_bytes = build_csv_1c_bytes([sample_receipt_result])

    # UTF-8 BOM for utf-8-sig
    assert csv_bytes.startswith(b"\xef\xbb\xbf")

    text = csv_bytes.decode("utf-8-sig")
    first_line = text.splitlines()[0]
    assert first_line == ";".join(EXPECTED_HEADERS)
    assert ";" in text
    assert "Распознано из фото чека" in text


def test_excel_1c_empty_items_creates_manual_check_row(tmp_path, sample_receipt_result):
    sample_receipt_result["items"] = []
    out_path = tmp_path / "out_empty_items.xlsx"
    build_excel_1c([sample_receipt_result], str(out_path))

    wb = load_workbook(out_path)
    ws_check = wb["Проверка"]
    ws_import = wb["1C_Импорт"]

    assert ws_check["B11"].value == "Проверить вручную"
    assert ws_import["F2"].value == "Нет данных"


def test_csv_1c_empty_items_creates_fallback_row(sample_receipt_result):
    sample_receipt_result["items"] = []
    text = build_csv_1c_bytes([sample_receipt_result]).decode("utf-8-sig")

    assert "Нет данных" in text


def test_excel_1c_skips_items_without_name(tmp_path, sample_receipt_result):
    sample_receipt_result["items"] = [
        {"name": "Товар 1", "quantity": 1, "price_per_unit": 100, "total_price": 100},
        {"name": "", "quantity": None, "price_per_unit": None, "total_price": None, "vat_rate": "20%", "vat_amount": 20},
        {"name": "   ", "quantity": None, "price_per_unit": None, "total_price": None, "vat_rate": "10%", "vat_amount": 10},
    ]
    out_path = tmp_path / "out_skip_empty_names.xlsx"
    build_excel_1c([sample_receipt_result], str(out_path))

    wb = load_workbook(out_path)
    ws_import = wb["1C_Импорт"]
    assert ws_import.max_row == 2
    assert ws_import["F2"].value == "Товар 1"


def test_csv_1c_skips_items_without_name(sample_receipt_result):
    sample_receipt_result["items"] = [
        {"name": "Товар 1", "quantity": 1, "price_per_unit": 100, "total_price": 100},
        {"name": "", "quantity": None, "price_per_unit": None, "total_price": None, "vat_rate": "20%", "vat_amount": 20},
    ]
    text = build_csv_1c_bytes([sample_receipt_result]).decode("utf-8-sig")
    lines = text.splitlines()
    assert len(lines) == 2
    assert "Товар 1" in lines[1]
