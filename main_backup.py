import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from src.vision_utils import prepare_image
from src.openai_client import extract_receipt_data_from_image


def process_receipt(image_path):
    """Обрабатывает одно изображение через OpenAI Vision API"""
    print(f"\n🔄 Обработка {image_path}...")

    try:
        prepared_path = prepare_image(image_path)
    except Exception as e:
        print(f"❌ Ошибка при подготовке изображения: {e}")
        return None

    data = extract_receipt_data_from_image(prepared_path)
    return data


def auto_fit_columns(worksheet):
    """Автоматически подбирает ширину колонок по содержимому"""
    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def save_to_excel(results, filename="receipts.xlsx"):
    """Сохраняет результаты в Excel-файл с двумя листами: сводка и товары."""
    wb = Workbook()

    # --- Лист 1: Сводка по чекам ---
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    summary_headers = ["№", "Номер чека", "Организация", "ИНН", "Дата", "Общая сумма", "НДС всего"]
    ws_summary.append(summary_headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, r in enumerate(results, 1):
        receipt = r.get("receipt", {}) or {}
        merchant = r.get("merchant", {}) or {}
        totals = r.get("totals", {}) or {}
        taxes = r.get("taxes", {}) or {}

        receipt_number = receipt.get("receipt_number", "")
        organization = merchant.get("organization", "")
        inn = merchant.get("inn", "")
        date = receipt.get("date", "")
        total = totals.get("total", "")
        total_vat = taxes.get("total_vat", "")

        row = [
            idx,
            receipt_number,
            organization,
            inn,
            date,
            total,
            total_vat,
        ]
        ws_summary.append(row)

    auto_fit_columns(ws_summary)

    # --- Лист 2: Товары (детализация) ---
    ws_items = wb.create_sheet("Товары")

    items_headers = ["№", "Номер чека", "Организация", "ИНН", "Дата",
                     "Наименование товара", "Цена за ед.", "Кол-во",
                     "Стоимость", "Ставка НДС", "Сумма НДС"]
    ws_items.append(items_headers)

    for col, header in enumerate(items_headers, 1):
        cell = ws_items.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, r in enumerate(results, 1):
        receipt = r.get("receipt", {}) or {}
        merchant = r.get("merchant", {}) or {}
        items = r.get("items", []) or []

        organization = merchant.get("organization", "") or ""
        inn = merchant.get("inn", "") or ""
        date = receipt.get("date", "") or ""
        receipt_number = receipt.get("receipt_number", "") or ""
        if not items:
            ws_items.append([idx, receipt_number, organization, inn, date, "Нет данных", "", "", "", "", ""])
        else:
            for item in items:
                row = [
                    idx,
                    receipt_number,
                    organization,
                    inn,
                    date,
                    item.get("name", ""),
                    item.get("price_per_unit", ""),
                    item.get("quantity", ""),
                    item.get("total_price", ""),
                    item.get("vat_rate", ""),
                    item.get("vat_amount", "")
                ]
                ws_items.append(row)

    auto_fit_columns(ws_items)

    wb.save(filename)
    print(f"\n✅ Результаты сохранены в файл {filename}")


def main():
    if len(sys.argv) < 2:
        print("Использование: python main.py <путь_к_изображению_или_папке>")
        return

    path = sys.argv[1]

    if os.path.isfile(path):
        files = [path]
    else:
        files = [os.path.join(path, f) for f in os.listdir(path)
                 if f.lower().endswith(('.jpg', '.jpeg', '.png', '.heic', '.heif'))]

    if not files:
        print(f"❌ Не найдено изображений (.jpg, .jpeg, .png, .heic, .heif) по пути: {path}")
        print("💡 Убедитесь, что папка содержит файлы с расширением .jpg, .jpeg, .png, .heic или .heif")
        return

    results = []
    for f in files:
        res = process_receipt(f)
        if res:
            results.append(res)
            print(f"✅ Чек обработан успешно")
        else:
            print(f"❌ Не удалось обработать чек")

    if not results:
        print("❌ Нет данных для сохранения.")
        return

    for i, r in enumerate(results, 1):
        receipt = r.get("receipt", {}) or {}
        merchant = r.get("merchant", {}) or {}
        totals = r.get("totals", {}) or {}
        taxes = r.get("taxes", {}) or {}
        receipt_number = receipt.get("receipt_number", "не указан")

        print(f"\n📄 --- Чек {i} (Номер на чеке: {receipt_number}) ---")
        print(f"Организация: {merchant.get('organization')}")
        print(f"ИНН: {merchant.get('inn')}")
        print(f"Дата: {receipt.get('date')}")
        print(f"Всего: {totals.get('total')} руб.")
        print(f"НДС всего: {taxes.get('total_vat')} руб.")
        print("Товары:")
        for item in r.get('items', []):
            name = item.get('name', '')
            price = item.get('total_price', '')
            print(f"  {name}: {price} руб.")

    save_to_excel(results, "receipts.xlsx")


if __name__ == "__main__":
    main()
